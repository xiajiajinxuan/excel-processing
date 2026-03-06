import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import shutil
import yaml

def process(data_df, **kwargs):
    """
    处理连续工作超过6天的数据
    :param data_df: 输入的DataFrame
    :param kwargs: 包含excel_file参数
    :return: 处理后的DataFrame
    """
    # 获取Excel文件路径
    excel_file = kwargs.get('excel_file')
    if not excel_file:
        raise ValueError("未提供Excel文件路径")

    # 获取项目根目录，配置文件位于 config/config.yaml
    project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    config_path = os.path.join(project_root, 'config', 'config.yaml')
    with open(config_path, 'r', encoding='utf-8') as f:
        config = yaml.safe_load(f)
    
    # 获取当前规则对应的模板文件名
    template_name = None
    for rule_id, rule_info in config.get('rules', {}).items():
        if rule_id == 'continuous_work_rule':
            template_name = rule_info.get('template')
            break
    
    if not template_name:
        raise ValueError("未找到对应的模板文件配置")

    # 获取模板文件路径
    template_dir = os.path.join(os.path.dirname(os.path.dirname(excel_file)), 'templates')
    template_file = os.path.join(template_dir, template_name)

    if not os.path.exists(template_file):
        raise ValueError(f"模板文件不存在: {template_file}")

    # 加载工作簿
    wb = load_workbook(template_file)
    # 优先选择"结果"Sheet
    if "结果" in wb.sheetnames:
        ws = wb["结果"]
    else:
        ws = wb.active

    # 定义红色填充
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    # 打印表头信息
    print("\n表头信息:")
    for col in range(1, ws.max_column + 1):
        print(f"列 {col}: {ws.cell(row=1, column=col).value}")

    # 从第三行开始处理数据
    for row in range(3, ws.max_row + 1):
        print(f"\n处理第 {row} 行:")
        # 获取该行的工作记录（从第二列开始）
        work_days = []
        for col in range(2, ws.max_column + 1):
            cell_value = ws.cell(row=row, column=col).value
            print(f"列 {col} 的值: {cell_value}, 类型: {type(cell_value)}")
            # 检查单元格值是否为数字且大于0，或者是否为日期格式
            is_work = False
            if isinstance(cell_value, (int, float)):
                is_work = cell_value > 0
            elif isinstance(cell_value, datetime):
                is_work = True
            work_days.append((col, is_work))

        # 计算连续工作天数并标记
        current_consecutive = 0
        start_col = None
        consecutive_ranges = []  # 存储所有连续工作超过6天的范围

        for col, is_work in work_days:
            if is_work:
                current_consecutive += 1
                if start_col is None:
                    start_col = col
            else:
                if current_consecutive > 6:
                    consecutive_ranges.append((start_col, col))
                current_consecutive = 0
                start_col = None

        # 处理行末的连续工作天数
        if current_consecutive > 6:
            consecutive_ranges.append((start_col, ws.max_column + 1))

        print(f"连续工作范围: {consecutive_ranges}")

        # 标记所有连续工作超过6天的范围
        for start, end in consecutive_ranges:
            # 确保end不超过最大列数
            end = min(end, ws.max_column + 1)
            for col in range(start, end):
                cell = ws.cell(row=row, column=col)
                cell.fill = red_fill

    # 确保output目录存在
    output_dir = os.path.join(os.path.dirname(os.path.dirname(excel_file)), 'output')
    os.makedirs(output_dir, exist_ok=True)

    # 生成输出文件名
    file_name = os.path.basename(excel_file)
    name_without_ext = os.path.splitext(file_name)[0]
    output_file = os.path.join(output_dir, f"{name_without_ext}_processed.xlsx")

    try:
        # 保存修改后的文件
        wb.save(output_file)
    except PermissionError:
        raise PermissionError(f"无法保存文件，请确保文件 {output_file} 未被其他程序打开")

    return data_df

def get_rule_info():
    """
    返回规则信息
    """
    return {
        "name": "连续工作超时检测",
        "description": "检测连续工作超过6天的记录并标记对应单元格为红色",
        "version": "1.0",
        "author": "System"
    } 